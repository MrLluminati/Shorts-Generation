from __future__ import annotations

import csv
from datetime import date, timedelta
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
IMPL = ROOT / "implementation"
OUTPUTS = ROOT / "outputs"

START_DATE = date(2026, 6, 8)
BASELINE = {
    "subscribers": 14,
    "valid_watch_hours": 0,
    "valid_shorts_views": 9,
    "total_public_views": 37854,
    "videos": 19,
}


SOURCE_LINKS = {
    "YouTube Partner Program eligibility": "https://support.google.com/youtube/answer/72851?hl=en",
    "YouTube reused content policy": "https://support.google.com/youtube/answer/1311392?hl=en-EN",
    "YouTube monetizable content guidance": "https://support.google.com/youtube/answer/2490020?hl=en-EN",
}


FORMATS = [
    {
        "name": "Why this scene works",
        "beat1": "State the surface-level scene in one short sentence.",
        "beat2": "Reveal the emotional or story reason the scene hits.",
        "beat3": "End with the consequence for the character or franchise.",
    },
    {
        "name": "Hidden detail you missed",
        "beat1": "Freeze or zoom on the exact detail viewers missed.",
        "beat2": "Explain what the detail silently tells us.",
        "beat3": "Connect it to a later payoff or theme.",
    },
    {
        "name": "Character decision explained",
        "beat1": "Show the impossible choice in plain words.",
        "beat2": "Explain what the character actually wants.",
        "beat3": "Explain why the decision changes the story.",
    },
]


TOPICS = [
    {
        "pillar": "Chainsaw Man Reze Arc",
        "concept": "Why Reze's smile is scarier than the explosion",
        "hook": "THIS SMILE WAS THE REAL WARNING",
        "title": "Reze's Smile Was the Warning | Chainsaw Man Explained",
        "tags": "#chainsawman #reze #animeexplained #shorts",
    },
    {
        "pillar": "Chainsaw Man Reze Arc",
        "concept": "Why Denji still wants someone to choose his heart",
        "hook": "DENJI WAS NOT TALKING ABOUT POWER",
        "title": "Denji's Heart Line Hits Different | Chainsaw Man Explained",
        "tags": "#denji #chainsawman #animeedit #shorts",
    },
    {
        "pillar": "Chainsaw Man Reze Arc",
        "concept": "City mouse versus country mouse as the movie's hidden question",
        "hook": "THIS FABLE EXPLAINS THE WHOLE MOVIE",
        "title": "City Mouse vs Country Mouse | Reze Arc Explained",
        "tags": "#chainsawman #reze #animeexplained #shorts",
    },
    {
        "pillar": "One Piece",
        "concept": "Why Whitebeard's final words still shake the world",
        "hook": "ONE LINE CHANGED THE PIRATE WORLD",
        "title": "Whitebeard's Final Prophecy Explained | One Piece",
        "tags": "#onepiece #whitebeard #animeexplained #shorts",
    },
    {
        "pillar": "One Piece",
        "concept": "Why Luffy's shock moments work better when everyone else reacts first",
        "hook": "LUFFY'S POWER IS IN THE REACTION",
        "title": "Why Everyone Reacting Makes Luffy Stronger | One Piece",
        "tags": "#luffy #onepiece #animeedit #shorts",
    },
    {
        "pillar": "Demon Slayer Infinity Castle",
        "concept": "Why spectacle alone is not the reason Infinity Castle works",
        "hook": "THE CASTLE IS NOT JUST A COOL BACKGROUND",
        "title": "Infinity Castle's Real Trick | Demon Slayer Explained",
        "tags": "#demonslayer #infinitycastle #animeexplained #shorts",
    },
    {
        "pillar": "Stranger Things",
        "concept": "Why music became the perfect weapon against Vecna",
        "hook": "THE MUSIC WAS NOT JUST NOSTALGIA",
        "title": "Why Music Works Against Vecna | Stranger Things Explained",
        "tags": "#strangerthings #vecna #seriesexplained #shorts",
    },
    {
        "pillar": "Stranger Things",
        "concept": "Why Will's quiet scenes matter more than action scenes",
        "hook": "WILL'S SILENCE IS THE CLUE",
        "title": "Why Will's Quiet Scenes Matter | Stranger Things",
        "tags": "#strangerthings #willbyers #netflix #shorts",
    },
    {
        "pillar": "The Dark Knight",
        "concept": "Why Joker's hospital pause makes the explosion funnier and scarier",
        "hook": "THE PAUSE MADE THIS SCENE ICONIC",
        "title": "The Joker Hospital Pause Explained | The Dark Knight",
        "tags": "#thedarkknight #joker #movieexplained #shorts",
    },
    {
        "pillar": "Batman / DC",
        "concept": "Why Batman scenes work best when the villain owns the frame",
        "hook": "BATMAN IS NOT ALWAYS THE CENTER",
        "title": "Why Villains Steal Batman Scenes | Movie Explained",
        "tags": "#batman #dc #movieexplained #shorts",
    },
    {
        "pillar": "Gachiakuta",
        "concept": "Why the Watchman reveal works as a mystery payoff",
        "hook": "THE REVEAL WORKED BECAUSE OF ONE SETUP",
        "title": "Watchman Reveal Explained | Gachiakuta",
        "tags": "#gachiakuta #animeexplained #anime #shorts",
    },
    {
        "pillar": "June/July 2026 Movies",
        "concept": "Use an official trailer moment and explain the one shot that sells the movie",
        "hook": "THIS TRAILER SHOT IS DOING ALL THE WORK",
        "title": "The Trailer Shot That Sells the Movie | Movie Explained",
        "tags": "#movieexplained #trailers #cinema #shorts",
    },
]


LAUNCH_SHORTS = [
    {
        "file": "03_denji_heart.mp4",
        "pillar": "Chainsaw Man Reze Arc",
        "concept": "Why Denji's heart line is funny and sad at the same time",
        "hook": "DENJI WAS NOT TALKING ABOUT POWER",
        "title": "Denji's Heart Line Hits Different | Chainsaw Man Explained",
        "tags": "#denji #chainsawman #animeexplained #shorts",
    },
    {
        "file": "01_date_to_devil.mp4",
        "pillar": "Chainsaw Man Reze Arc",
        "concept": "Why Reze's smile is scarier than the explosion",
        "hook": "THIS SMILE WAS THE REAL WARNING",
        "title": "Reze's Smile Was the Warning | Chainsaw Man Explained",
        "tags": "#chainsawman #reze #animeexplained #shorts",
    },
    {
        "file": "02_bomb_walked_in.mp4",
        "pillar": "Chainsaw Man Reze Arc",
        "concept": "Why the Bomb Devil entrance works before the fight starts",
        "hook": "THE DANGER ARRIVED QUIETLY FIRST",
        "title": "Bomb Devil Walked In Smiling | Chainsaw Man Explained",
        "tags": "#bombdevil #chainsawman #animeexplained #shorts",
    },
    {
        "file": "06_city_mouse.mp4",
        "pillar": "Chainsaw Man Reze Arc",
        "concept": "City mouse versus country mouse as the movie's hidden question",
        "hook": "THIS FABLE EXPLAINS THE WHOLE MOVIE",
        "title": "City Mouse vs Country Mouse | Reze Arc Explained",
        "tags": "#chainsawman #reze #animeexplained #shorts",
    },
    {
        "file": "04_easy_choice.mp4",
        "pillar": "Chainsaw Man Reze Arc",
        "concept": "Why Denji's easy decision is not actually simple",
        "hook": "KILL HER OR LET HER KILL YOU",
        "title": "Chainsaw Man Had One Brutal Choice | Reze Arc Explained",
        "tags": "#chainsawman #reze #animefight #shorts",
    },
    {
        "file": "05_run_away_together.mp4",
        "pillar": "Chainsaw Man Reze Arc",
        "concept": "Why Denji still asks Reze to run away after everything",
        "hook": "HE STILL ASKED HER TO RUN AWAY",
        "title": "He Still Asked Reze to Run Away | Chainsaw Man Explained",
        "tags": "#denji #reze #chainsawman #shorts",
    },
]


def write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content.strip() + "\n", encoding="utf-8")


def build_calendar() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for item_index in range(90):
        day_index = item_index // 3
        slot = item_index % 3
        upload_date = START_DATE + timedelta(days=day_index)
        topic = LAUNCH_SHORTS[item_index] if item_index < len(LAUNCH_SHORTS) else TOPICS[(item_index - len(LAUNCH_SHORTS)) % len(TOPICS)]
        fmt = FORMATS[(day_index + slot) % len(FORMATS)]
        rows.append(
            {
                "Day": str(day_index + 1),
                "Date": upload_date.isoformat(),
                "Slot": ["A", "B", "C"][slot],
                "Asset File": topic.get("file", "To create"),
                "Pillar": topic["pillar"],
                "Format": fmt["name"],
                "Concept": topic["concept"],
                "Opening Hook Text": topic["hook"],
                "On-screen Beat 1": fmt["beat1"],
                "On-screen Beat 2": fmt["beat2"],
                "On-screen Beat 3": fmt["beat3"],
                "Pinned Comment": "Which scene should I explain next?",
                "Upload Title": topic["title"],
                "Description": (
                    f"Silent explainer edit: {topic['concept']}. "
                    "This Short adds original context and scene breakdown text."
                ),
                "Hashtags": topic["tags"],
                "Production Notes": (
                    "Use only footage you have rights to, YouTube Remix-allowed clips, "
                    "or official trailer/clip snippets. Add visible original commentary text."
                ),
            }
        )
    return rows


def write_calendar(rows: list[dict[str, str]]) -> None:
    IMPL.mkdir(parents=True, exist_ok=True)
    csv_path = IMPL / "30_day_shorts_calendar.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def build_tracker(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Upload Tracker"

    headers = [
        "Day",
        "Date",
        "Slot",
        "Asset File",
        "Pillar",
        "Format",
        "Upload Title",
        "Video URL",
        "Published?",
        "24h Views",
        "72h Views",
        "7d Views",
        "Subscribers Gained",
        "Likes",
        "Comments",
        "Avg % Viewed",
        "Viewed vs Swiped %",
        "Likes / 1K Views",
        "Subs / 10K Views",
        "Decision",
        "Notes",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                int(row["Day"]),
                row["Date"],
                row["Slot"],
                row["Asset File"],
                row["Pillar"],
                row["Format"],
                row["Upload Title"],
                "",
                "No",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    for row_num in range(2, len(rows) + 2):
        ws[f"R{row_num}"] = f'=IF(L{row_num}>0,N{row_num}/L{row_num}*1000,"")'
        ws[f"S{row_num}"] = f'=IF(L{row_num}>0,M{row_num}/L{row_num}*10000,"")'
        ws[f"T{row_num}"] = (
            f'=IF(AND(L{row_num}>=1000,P{row_num}>=70),"Double down",'
            f'IF(AND(L{row_num}<100,P{row_num}<>""),"Retire hook","Wait"))'
        )

    dash = wb.create_sheet("Dashboard")
    dash_rows = [
        ["Metric", "Current", "Goal", "Gap", "Notes"],
        ["Subscribers", BASELINE["subscribers"], 1000, "=C2-B2", "Full YPP ads requirement"],
        ["Valid public watch hours", BASELINE["valid_watch_hours"], 4000, "=C3-B3", "Shorts Feed watch time does not count"],
        ["Valid public Shorts views", BASELINE["valid_shorts_views"], 10000000, "=C4-B4", "Alternative full YPP path"],
        ["Total public channel views", BASELINE["total_public_views"], "", "", "Public channel page"],
        ["Videos", BASELINE["videos"], "", "", "Public channel page"],
        ["30-day uploads planned", len(rows), "", "", "3 Shorts/day"],
        ["Published uploads", '=COUNTIF(\'Upload Tracker\'!H:H,"Yes")', "", "", "Update after publishing"],
        ["Total 7d views from sprint", "=SUM('Upload Tracker'!K:K)", "", "", "Fill tracker after 7 days"],
        ["Total subscribers gained", "=SUM('Upload Tracker'!L:L)", "", "", "Fill tracker from Studio"],
    ]
    for r in dash_rows:
        dash.append(r)

    gate = wb.create_sheet("Quality Gate")
    gate_rows = [
        ["Question", "Pass requirement"],
        ["Is the first frame understandable without audio?", "Yes"],
        ["Does the first 2 seconds include a specific hook?", "Yes"],
        ["Is there original commentary, not only subtitles?", "Yes"],
        ["Can a reviewer tell what you added?", "Yes"],
        ["Is the Short 25-45 seconds unless the story needs more?", "Preferred"],
        ["Are hashtags limited to 3-5 relevant tags?", "Yes"],
        ["Is there a pinned comment question?", "Yes"],
        ["Does the title include the franchise and explanation promise?", "Yes"],
    ]
    for r in gate_rows:
        gate.append(r)

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for col in range(1, sheet.max_column + 1):
            width = 16
            if col in [4, 5, 6, 7, 20]:
                width = 28
            sheet.column_dimensions[get_column_letter(col)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(IMPL / "monetization_sprint_tracker.xlsx")


def channel_profile_doc() -> str:
    return f"""
# Channel Profile Implementation

## Public Positioning
- Recommended channel name: **SceneCipher HQ**
- Recommended handle to try first: **@scenecipherhq**
- Backup handles to try if YouTube Studio rejects it: **@scenecipherlab**, **@plotcipherhq**, **@theframecipher**
- Brand meaning: **SceneCipher decodes scenes.** The channel is not just posting clips; it reveals the hidden clue, character choice, or emotional trick inside a scene.
- Tagline: **Decode the scene before it ends.**
- Channel promise: **Silent anime and movie explainers that decode hidden details, scene context, and character choices.**

## YouTube Studio Description
Use this as the channel description:

```text
Decode the scene before it ends.

SceneCipher HQ breaks down iconic anime and movie moments through silent scene explainers, hidden-detail breakdowns, character choices, trailer clues, and emotional payoffs.

What you will find here:
- Silent scene explainers with original on-screen commentary
- Hidden details you may have missed
- Character decisions explained
- Anime and movie moments with context, not random clip dumps

Current focus: Chainsaw Man, One Piece, Demon Slayer, Stranger Things, Batman/DC, and trending cinema moments.

For copyright or content removal requests, contact: abhikr14118@gmail.com
```

## Banner Text
```text
SCENECIPHER HQ
Decode the scene before it ends.
```

## About Tab Links/Labels
- Business email: keep existing email.
- First visible link label, if you add one later: **Request a Scene**
- Community post prompt: **Which anime or movie scene should I explain next?**

## Brand Voice
- Use words like **decode**, **hidden clue**, **scene logic**, **why it works**, **the real reason**, and **what changed here**.
- Avoid generic phrases like **best movie clips**, **viral clips**, **epic scenes**, and **must watch** as the main promise.
- Treat each Short as a mini answer, not a repost.

## Live Change Checklist
1. YouTube Studio -> Customization -> Profile: update the **Name** from "Absolute Cinema 2026" to **SceneCipher HQ**.
2. Try the **Handle** `@scenecipherhq`. If unavailable, try `@scenecipherlab`, then `@plotcipherhq`, then `@theframecipher`.
3. YouTube Studio -> Customization -> Basic info: paste the description above.
4. YouTube Studio -> Customization -> Branding: update banner text to match the new promise.
5. YouTube Studio -> Layout: feature a Shorts shelf first and pin the best-performing explainer Short once it has clear traction.
"""


def production_sop_doc() -> str:
    return """
# Silent Explainer Production SOP

## Non-negotiable Upload Rules
- Publish **3 Shorts/day** for 30 days.
- Each Short must include original on-screen commentary, not only subtitles or hashtags.
- Keep most Shorts **25-45 seconds**; never exceed 59 seconds during this sprint.
- Use 1080x1920 vertical MP4, H.264 video, AAC audio, and readable text inside mobile safe zones.
- Use only footage you have rights to, YouTube Remix-allowed clips, official trailer/clip snippets, or transformative excerpts with clear original value.

## 45-Second Structure
```text
0-2s: Big hook text
2-8s: What the viewer is seeing
8-22s: Hidden detail / emotional reason / story context
22-35s: Why it matters
35-45s: Payoff + question CTA
```

## Hook Bank
- THIS SMILE WAS THE REAL WARNING
- THIS FIGHT WAS NOT ABOUT POWER
- MOST PEOPLE MISSED THIS DETAIL
- THE PAUSE MADE THIS SCENE ICONIC
- THIS ONE SHOT EXPLAINS THE CHARACTER
- THE VILLAIN WON BEFORE THE FIGHT STARTED
- THIS LINE CHANGED THE WHOLE STORY
- THE TRAILER IS HIDING THE REAL CONFLICT

## Editing Checklist
- First frame must be readable without audio.
- Add one text idea at a time; no paragraph walls.
- Use zooms or freeze frames only when they help explain the point.
- Avoid generic "viral", "trending", and repeated hashtag stuffing.
- End with a specific question: "Which scene should I explain next?"

## Weekly Review Rhythm
- Every morning: record 24h metrics for yesterday's three uploads.
- Every third day: record 72h metrics and choose one hook to repeat.
- Every Sunday: retire the weakest format and create sequels for the top two Shorts.
"""


def existing_shorts_doc() -> str:
    return """
# Existing Chainsaw Man Shorts Launch Plan

Use the six rendered Shorts as the first two days of the sprint, but upload them with explainer positioning.

## Upload Order
1. `03_denji_heart.mp4` - strongest emotional/comedy hook.
2. `01_date_to_devil.mp4` - betrayal hook.
3. `02_bomb_walked_in.mp4` - threat reveal.
4. `06_city_mouse.mp4` - story-theme explainer.
5. `04_easy_choice.mp4` - stakes/choice explainer.
6. `05_run_away_together.mp4` - emotional payoff.

## Revised Metadata

### 03_denji_heart.mp4
- Title: Denji's Heart Line Hits Different | Chainsaw Man Explained #shorts
- Description: Denji is not just complaining here. This line explains the whole tragedy of people wanting Chainsaw's power instead of Denji himself.
- Pinned comment: Did Denji mean this as a joke, or was it actually sad?
- Hashtags: #denji #chainsawman #animeexplained #shorts

### 01_date_to_devil.mp4
- Title: Reze's Smile Was the Warning | Chainsaw Man Explained #shorts
- Description: The scary part is not the attack. It is how calm Reze is before the betrayal lands.
- Pinned comment: When did you realize Reze was dangerous?
- Hashtags: #chainsawman #reze #animeedit #shorts

### 02_bomb_walked_in.mp4
- Title: Bomb Devil Walked In Smiling | Chainsaw Man Explained #shorts
- Description: The scene works because everyone reacts before the fight starts. The danger arrives quietly first.
- Pinned comment: Is a quiet villain entrance better than a loud one?
- Hashtags: #bombdevil #chainsawman #animeexplained #shorts

### 06_city_mouse.mp4
- Title: The City Mouse Question Explains Reze | Chainsaw Man #shorts
- Description: The fable is not random. It explains what Reze wants and why Denji cannot understand the danger.
- Pinned comment: Are you choosing the city mouse or country mouse?
- Hashtags: #reze #chainsawman #animeexplained #shorts

### 04_easy_choice.mp4
- Title: Chainsaw Man Had One Brutal Choice | Reze Arc Explained #shorts
- Description: Angel gives Denji the clean moral answer, but Denji never works like a normal hero.
- Pinned comment: Was Denji's answer brave or foolish?
- Hashtags: #chainsawman #reze #animefight #shorts

### 05_run_away_together.mp4
- Title: He Still Asked Reze to Run Away | Chainsaw Man Explained #shorts
- Description: After everything Reze did, Denji's response shows exactly why this arc hurts.
- Pinned comment: Would Reze have gone with him if things were different?
- Hashtags: #denji #reze #chainsawman #shorts
"""


def metadata_doc() -> str:
    return """
# Shorts Metadata - Explainer Version

All six files are ready to upload as under-60-second vertical Shorts. Use the revised titles, descriptions, hashtags, and pinned comments from `implementation/existing_shorts_launch_plan.md`.

Recommended upload dates:
- Day 1: `03_denji_heart.mp4`, `01_date_to_devil.mp4`, `02_bomb_walked_in.mp4`
- Day 2: `06_city_mouse.mp4`, `04_easy_choice.mp4`, `05_run_away_together.mp4`

Do not upload all six at the exact same minute. Space them by at least 3-4 hours.
"""


def readme_doc() -> str:
    links = "\n".join(f"- [{name}]({url})" for name, url in SOURCE_LINKS.items())
    return f"""
# SceneCipher HQ Monetization Sprint Package

This folder implements the 30-day plan for rebranding Absolute Cinema 2026 into **SceneCipher HQ**.

## Baseline
- Subscribers: **{BASELINE['subscribers']} / 1,000**
- Valid public watch hours: **{BASELINE['valid_watch_hours']} / 4,000**
- Valid Shorts views: **{BASELINE['valid_shorts_views']} / 10,000,000**
- Public channel total views: **{BASELINE['total_public_views']}**
- Public videos: **{BASELINE['videos']}**

## Files
- `channel_profile_implementation.md`: exact channel copy and YouTube Studio checklist.
- `silent_explainer_sop.md`: production rules, hook structure, editing checklist.
- `30_day_shorts_calendar.csv`: 90 upload ideas for 3 Shorts/day.
- `monetization_sprint_tracker.xlsx`: tracker with dashboard and formulas.
- `existing_shorts_launch_plan.md`: upload order and metadata for the six rendered Shorts.

## Sources
{links}
"""


def write_docs(rows: list[dict[str, str]]) -> None:
    write(IMPL / "README.md", readme_doc())
    write(IMPL / "channel_profile_implementation.md", channel_profile_doc())
    write(IMPL / "silent_explainer_sop.md", production_sop_doc())
    write(IMPL / "existing_shorts_launch_plan.md", existing_shorts_doc())
    write(OUTPUTS / "shorts_metadata_explainer.md", metadata_doc())


def make_zip() -> None:
    zip_path = OUTPUTS / "chainsaw_man_reze_explainer_launch_pack.zip"
    members = [
        OUTPUTS / "01_date_to_devil.mp4",
        OUTPUTS / "02_bomb_walked_in.mp4",
        OUTPUTS / "03_denji_heart.mp4",
        OUTPUTS / "04_easy_choice.mp4",
        OUTPUTS / "05_run_away_together.mp4",
        OUTPUTS / "06_city_mouse.mp4",
        OUTPUTS / "shorts_metadata_explainer.md",
        IMPL / "existing_shorts_launch_plan.md",
    ]
    with ZipFile(zip_path, "w", ZIP_DEFLATED) as zf:
        for member in members:
            if member.exists():
                zf.write(member, member.name)


def main() -> None:
    rows = build_calendar()
    write_calendar(rows)
    build_tracker(rows)
    write_docs(rows)
    make_zip()
    print(f"Wrote implementation package to {IMPL}")


if __name__ == "__main__":
    main()
